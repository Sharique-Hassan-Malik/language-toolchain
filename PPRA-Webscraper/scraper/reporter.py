"""Excel report builder and keyword detector.

The keyword-matching logic is deliberately kept free of any heavy dependency, so
it can be unit-tested without pandas, openpyxl, Selenium or Gmail installed —
`pandas`/`openpyxl` are imported lazily inside `save_excel` only.
"""

from __future__ import annotations

import re

from scraper.config import KEYWORDS, WHOLE_WORD


# ---------------------------------------------------------------------------
# Keyword matching (pure, dependency-free, and unit-tested)
# ---------------------------------------------------------------------------

def _keyword_pattern(keyword: str, whole_word: bool) -> re.Pattern:
    body = re.escape(keyword)
    if whole_word:
        body = rf"\b{body}\b"
    return re.compile(body, re.IGNORECASE)


def matched_keywords(details: str, keywords: list[str], whole_word: bool = WHOLE_WORD) -> list[str]:
    """The subset of `keywords` that occur in `details` (case-insensitive)."""
    return [kw for kw in keywords if _keyword_pattern(kw, whole_word).search(details)]


def highlight(details: str, keywords: list[str], whole_word: bool = WHOLE_WORD) -> str:
    """Wrap each keyword occurrence in <b>…</b> for the HTML email body."""
    out = details
    for kw in keywords:
        out = _keyword_pattern(kw, whole_word).sub(r"<b>\g<0></b>", out)
    return out


def detect_keywords(
    tender_data: list[dict],
    keywords: list[str] = KEYWORDS,
    whole_word: bool = WHOLE_WORD,
) -> tuple[dict[str, list[str]], list[str]]:
    """Find tenders whose Tender Details contain any keyword.

    Returns ``(matched_map, html_messages)`` where ``matched_map`` maps a tender
    number to the keywords it matched, and ``html_messages`` is one HTML
    paragraph per matched tender for the email body.
    """
    matched_map: dict[str, list[str]] = {}
    html_messages: list[str] = []
    seen: set[str] = set()

    for tender in tender_data:
        tender_no = tender.get("Tender No", "")
        details   = tender.get("Tender Details", "")
        hits = matched_keywords(details, keywords, whole_word)

        if hits and tender_no not in seen:
            seen.add(tender_no)
            matched_map[tender_no] = hits
            html_messages.append(
                f"<p><b>Tender {tender_no}</b> matches "
                f"<b>{', '.join(hits)}</b><br>"
                f"------<br>{highlight(details, hits, whole_word)}<br>------</p>"
            )

    if matched_map:
        print(f"\n{len(matched_map)} tender(s) matched keywords {keywords}:\n")
        for msg in html_messages:
            print(re.sub(r"<[^>]+>", "", msg))     # strip tags for the console log
    else:
        print(f"\nNo tenders matched keywords {keywords}.")

    return matched_map, html_messages


# ---------------------------------------------------------------------------
# Excel generation (pandas/openpyxl imported lazily)
# ---------------------------------------------------------------------------

def save_excel(
    tender_data: list[dict],
    excel_path: str,
    matched_map: dict[str, list[str]] | None = None,
) -> None:
    """Write `tender_data` to a styled .xlsx: auto-width columns, wrap-text,
    proportional row heights, and bold Tender Details for matched tenders."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    matched_map = matched_map or {}

    df = pd.DataFrame(tender_data)
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    col_index = {str(cell.value): i for i, cell in enumerate(ws[1], start=1) if cell.value}
    tender_no_col = col_index.get("Tender No")
    details_col   = col_index.get("Tender Details")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if tender_no_col and details_col:
            tender_no = row[tender_no_col - 1].value
            if tender_no in matched_map:
                row[details_col - 1].font = Font(bold=True)

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width  = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[letter].width = min(width + 5, 100)

    for row in ws.iter_rows():
        max_lines = max(
            (str(c.value).count("\n") + len(str(c.value)) // 50 + 1 for c in row if c.value),
            default=1,
        )
        ws.row_dimensions[row[0].row].height = max_lines * 15

    wb.save(excel_path)
    print(f"Excel report saved to {excel_path}")
