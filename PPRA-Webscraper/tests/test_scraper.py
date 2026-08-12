"""Tests for the parts of the scraper that don't need a browser or Gmail.

The scraping and emailing themselves need a live PPRA session and OAuth
credentials, but the logic that historically caused bugs — keyword matching,
filename sanitising, the HTML/email body, the Excel styling — is pure and is
tested here. Selenium and the Google libraries are imported lazily by the
package, so this suite runs with only pandas/openpyxl (and the stdlib) present.
"""

import os
import sys

import pytest

# import the package from the repo root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from scraper.util import normalize_tender_no, pdf_filename          # noqa: E402
from scraper.reporter import detect_keywords, highlight, matched_keywords  # noqa: E402
from scraper.emailer import build_message                            # noqa: E402


# ---------------------------------------------------------------------------
# filename / tender-number helpers
# ---------------------------------------------------------------------------

def test_normalize_tender_no_strips_view_invoice():
    assert normalize_tender_no("TS123/2025  View Invoice") == "TS123/2025"
    assert normalize_tender_no("  A\n  B  ") == "A B"


def test_pdf_filename_sanitises_slashes():
    # tender numbers contain '/', which breaks os.rename if left unsanitised
    assert "/" not in pdf_filename("TS123/2025")
    assert pdf_filename("TS123/2025") == "TS123-2025.pdf"
    assert pdf_filename('a:b*c?"d') == "a-b-c--d.pdf"          # all illegal chars replaced
    assert pdf_filename("") == "tender.pdf"                     # never empty


def test_pdf_filename_keeps_extension():
    assert pdf_filename("X", ".PDF") == "X.PDF"


# ---------------------------------------------------------------------------
# keyword matching
# ---------------------------------------------------------------------------

def test_matched_keywords_is_case_insensitive():
    assert matched_keywords("Supply to STATE bank", ["Bank"]) == ["Bank"]
    assert matched_keywords("nothing here", ["Bank"]) == []


def test_whole_word_matching_avoids_substrings():
    # substring match (default) catches "Embankment"; whole-word does not
    assert matched_keywords("Embankment works", ["Bank"], whole_word=False) == ["Bank"]
    assert matched_keywords("Embankment works", ["Bank"], whole_word=True) == []
    assert matched_keywords("The Bank job", ["Bank"], whole_word=True) == ["Bank"]


def test_highlight_bolds_each_occurrence():
    out = highlight("Bank and bank", ["Bank"])
    assert out.count("<b>") == 2
    assert "<b>Bank</b>" in out and "<b>bank</b>" in out       # preserves original case


def test_detect_keywords_dedupes_and_builds_messages():
    data = [
        {"Tender No": "T1", "Tender Details": "University lab supply"},
        {"Tender No": "T1", "Tender Details": "University lab supply"},   # duplicate row
        {"Tender No": "T2", "Tender Details": "road works"},              # no match
        {"Tender No": "T3", "Tender Details": "Bank branch fit-out"},
    ]
    matched, messages = detect_keywords(data, ["Bank", "University"])
    assert set(matched) == {"T1", "T3"}                # T2 excluded, T1 not duplicated
    assert matched["T1"] == ["University"]
    assert len(messages) == 2
    assert "<b>Tender T3</b>" in "".join(messages)


# ---------------------------------------------------------------------------
# email message construction
# ---------------------------------------------------------------------------

def test_build_message_is_valid_and_skips_missing_files(tmp_path):
    import base64
    from email import message_from_bytes

    real = tmp_path / "report.xlsx"
    real.write_bytes(b"fake xlsx bytes")
    msg = build_message(
        sender="me@example.com", recipient="you@example.com",
        subject="Subject", html_body="<p>hi</p>",
        files=[str(real), str(tmp_path / "missing.pdf")],   # one exists, one does not
    )
    assert set(msg) == {"raw"}
    parsed = message_from_bytes(base64.urlsafe_b64decode(msg["raw"]))
    assert parsed["subject"] == "Subject"
    assert parsed["to"] == "you@example.com"
    parts = list(parsed.walk())
    # one text/html body + exactly one attachment (the missing file was skipped)
    assert any(p.get_content_type() == "text/html" for p in parts)
    attachments = [p for p in parts if p.get("Content-Disposition", "").startswith("attachment")]
    assert len(attachments) == 1
    assert 'filename="report.xlsx"' in attachments[0]["Content-Disposition"]


# ---------------------------------------------------------------------------
# Excel report  (needs pandas/openpyxl; skipped if absent)
# ---------------------------------------------------------------------------

def test_save_excel_writes_a_styled_workbook(tmp_path):
    pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from scraper.reporter import save_excel

    data = [
        {"Sr No": "1", "Tender No": "T1", "Tender Details": "Bank fit-out",
         "Advertisement Date": "2025-01-01", "Closing Date": "2025-02-01"},
        {"Sr No": "2", "Tender No": "T2", "Tender Details": "road works",
         "Advertisement Date": "2025-01-03", "Closing Date": "2025-02-03"},
    ]
    out = tmp_path / "out.xlsx"
    save_excel(data, str(out), matched_map={"T1": ["Bank"]})
    assert out.exists()

    ws = load_workbook(out).active
    headers = [c.value for c in ws[1]]
    assert headers == ["Sr No", "Tender No", "Tender Details", "Advertisement Date", "Closing Date"]
    details_col = headers.index("Tender Details") + 1
    # the matched tender (row 2) has bold Tender Details; the unmatched (row 3) does not
    assert ws.cell(row=2, column=details_col).font.bold is True
    assert not ws.cell(row=3, column=details_col).font.bold
